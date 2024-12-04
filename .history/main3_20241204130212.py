import openpyxl

# Cargar el archivo Excel
workbook = openpyxl.load_workbook('tu_archivo.xlsx')

# Iterar sobre cada hoja
for sheet in workbook.sheetnames:
    hoja = workbook[sheet]
    print(f'Procesando hoja: {hoja.title}')
    
    # Iterar sobre cada fila en la hoja
    for row in hoja.iter_rows(values_only=True):
        print(row)  # Aquí puedes agregar el procesamiento que necesites

# Guardar los cambios si es necesario
workbook.save('tu_archivo_modificado.xlsx')