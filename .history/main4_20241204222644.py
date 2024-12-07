import fitz  # PyMuPDF
import openpyxl
import os

# Carpeta con archivos PDF
pdf_folder = 'C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados'  # Cambia esta ruta a la ubicación de tu carpeta

import fitz  # PyMuPDF
import openpyxl
import os

# Carpeta con archivos PDF
pdf_folder = '/mnt/data/pdf_folder'  # Cambia esta ruta a la ubicación de tu carpeta

# Crear un archivo Excel para almacenar los datos combinados
workbook = openpyxl.Workbook()
sheet = workbook.active

# Encabezados para la tabla
headers = [
    "ID", "Sexo", "RBC-ERITROCITOS", "HGB-HEMOGLOBINA", "HCT-HEMATOCRITO", "MCV", "MCH", "MCHC", 
    "WBC-LEUCOCITOS", "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", 
    "EOS%-EOSINOFILOS", "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", 
    "MON#-MONOCITOS", "EOS#-EOSINOFILOS", "BAS#-BASOFILOS", "RDW-CV", "RDW-SD", 
    "PLT-PLAQUETAS", "MPV", "PDW"
]

# Escribir encabezados en Excel
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Función para buscar y extraer datos relevantes usando palabras y posiciones
def extract_data_by_words(page):
    """
    Extrae datos relevantes basados en palabras y posiciones horizontales.
    """
    words = page.get_text("words")  # Extraer todas las palabras con sus posiciones
    data = {key: "" for key in headers}  # Inicializar el diccionario de datos

    # Buscar el ID en el texto completo
    text = page.get_text("text")
    if "ID:" in text:
        # Extraer el ID después de "ID:"
        start_index = text.index("ID:") + len("ID:")
        end_index = text.find("\n", start_index)
        data["ID"] = text[start_index:end_index].strip()

    for i, word in enumerate(words):
        x0, y0, x1, y1, text, block_no, line_no, word_no = word
        # Filtrar los datos según las etiquetas clave
        if "Sexo:" in text:
            data["Sexo"] = text.split(":")[-1].strip()
        elif text in headers:  # Detectar encabezados como RBC-ERITROCITOS, etc.
            # Buscar el valor en la misma línea pero más a la derecha
            next_word = next((w[4] for w in words[i + 1:] if w[1] == y0 and w[0] > x1), None)
            if next_word:
                data[text] = next_word

    # Rellenar manualmente los valores faltantes de etiquetas específicas
    missing_keys = [
        "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", "EOS%-EOSINOFILOS", 
        "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", "MON#-MONOCITOS", 
        "EOS#-EOSINOFILOS", "BAS#-BASOFILOS"
    ]
    for key in missing_keys:
        # Buscar las etiquetas y sus valores asociados
        for j, word in enumerate(words):
            _, _, _, _, current_text, _, current_line, _ = word
            if key in current_text:  # Encontrar el encabezado en cuestión
                try:
                    value = words[j + 1][4]  # Tomar el siguiente valor en la misma línea
                    data[key] = value
                except IndexError:
                    data[key] = "FALTANTE"  # Marca como "FALTANTE" si no se encuentra el valor
    return data

# Iterar sobre cada archivo PDF en la carpeta
row_num = 2
for pdf_file in os.listdir(pdf_folder):
    if pdf_file.endswith('.pdf'):
        pdf_path = os.path.join(pdf_folder, pdf_file)
        pdf = fitz.open(pdf_path)

        # Iterar sobre cada página del PDF y extraer datos
        for page_num in range(len(pdf)):
            page = pdf.load_page(page_num)
            data = extract_data_by_words(page)  # Extraer datos relevantes
            if any(data.values()):  # Verificar que hay datos válidos
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=row_num, column=col_num, value=data.get(header, ""))
                row_num += 1

# Guardar el archivo Excel con todos los datos
output_file = '/mnt/data/resultados_completos.xlsx'
workbook.save(output_file)
print(f"Archivo Excel generado: {output_file}")
