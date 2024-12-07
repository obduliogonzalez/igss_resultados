import fitz  # PyMuPDF
import openpyxl
import os

# Ruta a la carpeta donde están los archivos PDF
pdf_folder = 'C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados/saas.pdf'

# Crear un nuevo archivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Encabezados para la tabla
headers = [
    "ID", "Sexo", "RBC-ERITROCITOS", "HGB-HEMOGLOBINA", "HCT-HEMATOCRITO", "MCV", "MCH", "MCHC", 
    "WBC-LEUCOCITOS", "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", 
    "EOS%-EOSINOFILOS", "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", 
    "MON#-MONOCITOS", "EOS#-EOSINOFILOS", "BAS#-BASOFILOS", "RDW-CV", "RDW-SD", 
    "PLT-PLAQUETAS", "MPV", "PDW"
]

# Escribir encabezados en Excel
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Función para buscar y extraer datos relevantes usando bloques de texto
def extract_data_by_blocks(page):
    """
    Extrae datos relevantes basados en bloques de texto.
    """
    blocks = page.get_text("blocks")  # Extraer bloques de texto
    data = {key: "" for key in headers}  # Inicializar el diccionario de datos

    for block in blocks:
        _, _, _, _, block_text = block  # Extraer el texto del bloque
        lines = block_text.split('\n')  # Dividir el texto en líneas

        for line in lines:
            # Verificar si la línea contiene la etiqueta y extraer el valor
            if "ID:" in line:
                data["ID"] = line.split("ID:")[-1].strip()
            elif "Sexo:" in line:
                data["Sexo"] = line.split("Sexo:")[-1].strip()
            elif any(header in line for header in headers):
                for header in headers:
                    if header in line:
                        try:
                            # Buscar el valor después de la etiqueta y tomar el primer valor válido
                            value = line.split(header)[-1].split()[0]  # Captura el valor hasta el siguiente espacio
                            data[header] = value
                        except IndexError:
                            data[header] = "FALTANTE"  # Marca como "FALTANTE" si no se encuentra el valor

    return data

# Iterar sobre cada archivo PDF en la carpeta
row_num = 2
for pdf_file in os.listdir(pdf_folder):
    if pdf_file.endswith('.pdf'):
        pdf_path = os.path.join(pdf_folder, pdf_file)
        pdf = fitz.open(pdf_path)

        # Iterar sobre cada página del PDF y extraer datos
        for page_num in range(len(pdf)):
            page = pdf.load_page(page_num)
            data = extract_data_by_blocks(page)  # Extraer datos relevantes
            if any(data.values()):  # Verificar que hay datos válidos
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=row_num, column=col_num, value=data.get(header, ""))
                row_num += 1

# Guardar el archivo Excel con todos los datos
output_file = 'resultados_completos.xlsx'
workbook.save(output_file)
print(f"Archivo Excel generado: {output_file}")
