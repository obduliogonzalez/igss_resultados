import fitz  # PyMuPDF
import openpyxl

# Cargar el archivo PDF
pdf_document = 'C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados/GARCIA_GAYTAN_OLGA_JUDITH_1.pdf'
pdf = fitz.open(pdf_document)

# Crear un nuevo archivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Encabezados para la tabla
headers = [
    "ID", "Sexo", "RBC-ERITROCITOS", "HGB-HEMOGLOBINA", "HCT-HEMATOCRITO", "MCV", "MCH", "MCHC", 
    "WBC-LEUCOCITOS", "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", 
    "EOS%-EOSINOFILOS", "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", 
    "MON#-MONOCITOS", "EOS#-EOSINOFILOS", "BAS#-BASOFILOS", "RDW-CV", "RDW-SD", 
    "PLT-PLAQUETAS", "MPV", "PDW"
]

# Escribir encabezados en Excel
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Función para buscar y extraer datos relevantes usando palabras y posiciones
def extract_data_by_words(page):
    """
    Extrae datos relevantes basados en palabras y posiciones horizontales.
    """
    words = page.get_text("words")  # Extraer todas las palabras con sus posiciones
    data = {key: "" for key in headers}  # Inicializar el diccionario de datos

    for i, word in enumerate(words):
        x0, y0, x1, y1, text, block_no, line_no, word_no = word
        # Filtrar los datos según las etiquetas clave
        if "ID:" in text:
            data["ID"] = text.split(":")[-1].strip()
        elif "Sexo:" in text:
            data["Sexo"] = text.split(":")[-1].strip()
        elif text in headers:  # Detectar encabezados como RBC-ERITROCITOS, etc.
            # Buscar el valor en la misma línea pero más a la derecha
            next_word = next((w[4] for w in words[i + 1:] if w[1] == y0 and w[0] > x1), None)
            if next_word:
                data[text] = next_word

    # Rellenar los valores faltantes manualmente si están en un patrón especial
    for i, word in enumerate(words):
        _, _, _, _, text, _, line_no, _ = word
        if line_no in ["FORMULA ROJA", "FORMULA BLANCA"]:  # Detectar secciones específicas
            # Buscar los valores correspondientes según el índice
            try:
                if "RBC-ERITROCITOS" in text:
                    data["RBC-ERITROCITOS"] = words[i + 1][4]
                elif "HGB-HEMOGLOBINA" in text:
                    data["HGB-HEMOGLOBINA"] = words[i + 1][4]
                elif "WBC-LEUCOCITOS" in text:
                    data["WBC-LEUCOCITOS"] = words[i + 1][4]
                elif "RDW-CV" in text:
                    data["RDW-CV"] = words[i + 1][4]
                elif "RDW-SD" in text:
                    data["RDW-SD"] = words[i + 1][4]
                elif "PLT-PLAQUETAS" in text:
                    data["PLT-PLAQUETAS"] = words[i + 1][4]
                elif "MPV" in text:
                    data["MPV"] = words[i + 1][4]
                elif "PDW" in text:
                    data["PDW"] = words[i + 1][4]
            except IndexError:
                continue

    return data

# Iterar sobre cada página del PDF y extraer datos
row_num = 2
for page_num in range(len(pdf)):
    page = pdf.load_page(page_num)
    data = extract_data_by_words(page)  # Extraer datos relevantes
    if any(data.values()):  # Verificar que hay datos válidos
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num, value=data.get(header, ""))
        row_num += 1

# Guardar el archivo Excel
output_file = '/mnt/data/resultados_procesados_completos.xlsx'
workbook.save(output_file)
print(f"Archivo Excel generado: {output_file}")