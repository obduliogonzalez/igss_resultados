import openpyxl

# Cargar el archivo Excel
workbook = openpyxl.load_workbook("C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados/GARCIA_GAYTAN_OLGA_JUDITH_M.pdf")

import fitz  # PyMuPDF
import openpyxl

# Cargar el archivo PDF
pdf_document = 'tu_archivo.pdf'
pdf = fitz.open(pdf_document)

# Crear un nuevo archivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Iterar sobre cada página del PDF
for page_num in range(len(pdf)):
    page = pdf.load_page(page_num)
    text = page.get_text("text")
    
    # Dividir el texto en líneas y escribir cada línea en una fila de Excel
    for line_num, line in enumerate(text.split('\n')):
        sheet.cell(row=line_num + 1, column=page_num + 1, value=line)

# Guardar el archivo Excel
workbook.save('tu_archivo.xlsx')