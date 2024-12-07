import fitz  # PyMuPDF
import openpyxl

# Cargar el archivo PDF
pdf_document = 'C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados/GARCIA_GAYTAN_OLGA_JUDITH_M.pdf'
pdf = fitz.open(pdf_document)

# Crear un nuevo archivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Encabezados para la tabla (modifica según tu necesidad)
headers = [
    "ID", "Sexo", "RBC-ERITROCITOS", "HGB-HEMOGLOBINA", "HCT-HEMATOCRITO", "MCV", "MCH", "MCHC", 
    "WBC-LEUCOCITOS", "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", 
    "EOS%-EOSINOFILOS", "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", 
    "MON#-MONOCITOS", "EOS#-EOSINOFILOS", "BAS#-BASOFILOS", "RDW-CV", "RDW-SD", 
    "PLT-PLAQUETAS", "MPV", "PDW"
]

# Escribir encabezados en Excel
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Iterar sobre cada página del PDF
row_num = 2  # Comienza en la segunda fila porque la primera tiene los encabezados
for page_num in range(len(pdf)):
    page = pdf.load_page(page_num)
    text = page.get_text("blocks")
    
    # Depurar el contenido extraído
    print(f'Contenido de la página {page_num + 1}:\n{text}\n')
    
    # Procesar el contenido para extraer datos
    lines = text.split('\n')
    for line in lines:
        # Intentar dividir el texto por tabulaciones, espacios u otro delimitador
        columns = line.split()  # Ajusta según el formato de tu PDF
        if len(columns) == len(headers):  # Asegúrate de que coincida con el número de columnas esperadas
            for col_num, value in enumerate(columns, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
            row_num += 1

# Guardar el archivo Excel
workbook.save('resultados_procesados.xlsx')
