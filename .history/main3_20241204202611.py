import fitz  # PyMuPDF
import openpyxl

# Cargar el archivo PDF
pdf_document = 'C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados/GARCIA_GAYTAN_OLGA_JUDITH_M.pdf'
pdf = fitz.open(pdf_document)

# Crear un nuevo archivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Encabezados para la tabla
headers = [
    "ID", "Sexo", "RBC-ERITROCITOS", "HGB-HEMOGLOBINA", "HCT-HEMATOCRITO", "MCV", "MCH", "MCHC", 
    "WBC-LEUCOCITOS", "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", 
    "EOS%-EOSINOFILOS", "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", 
    "MON#-MONOCITOS", "EOS#-EOSINOFILOS", "BAS#-BASOFILOS", "RDW-CV", "RDW-SD", 
    "PLT-PLAQUETAS", "MPV", "PDW"
]

# Escribir encabezados en Excel
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Función para buscar y extraer datos relevantes
def extract_data(text):
    """
    Busca las líneas clave en el texto y extrae los datos.
    """
    data = {}
    for line in text.split('\n'):
        if "ID:" in line:
            data["ID"] = line.split(":")[-1].strip()
        elif "Sexo:" in line:
            data["Sexo"] = line.split(":")[-1].strip()
        elif "RBC-ERITROCITOS" in line:
            data["RBC-ERITROCITOS"] = line.split()[1]
        elif "HGB-HEMOGLOBINA" in line:
            data["HGB-HEMOGLOBINA"] = line.split()[1]
        elif "HCT-HEMATOCRITO" in line:
            data["HCT-HEMATOCRITO"] = line.split()[1]
        elif "MCV" in line:
            data["MCV"] = line.split()[1]
        elif "MCH" in line:
            data["MCH"] = line.split()[1]
        elif "MCHC" in line:
            data["MCHC"] = line.split()[1]
        elif "WBC-LEUCOCITOS" in line:
            data["WBC-LEUCOCITOS"] = line.split()[1]
        elif "LYM%-LINFOCITOS" in line:
            data["LYM%-LINFOCITOS"] = line.split()[1]
        elif "NEU%-NEUTROFILOS" in line:
            data["NEU%-NEUTROFILOS"] = line.split()[1]
        elif "MON%-MONOCITOS" in line:
            data["MON%-MONOCITOS"] = line.split()[1]
        elif "EOS%-EOSINOFILOS" in line:
            data["EOS%-EOSINOFILOS"] = line.split()[1]
        elif "BAS%-BASOFILOS" in line:
            data["BAS%-BASOFILOS"] = line.split()[1]
        elif "LYN#-LINFOCITOS" in line:
            data["LYN#-LINFOCITOS"] = line.split()[1]
        elif "NEU#-NEUTROFILOS" in line:
            data["NEU#-NEUTROFILOS"] = line.split()[1]
        elif "MON#-MONOCITOS" in line:
            data["MON#-MONOCITOS"] = line.split()[1]
        elif "EOS#-EOSINOFILOS" in line:
            data["EOS#-EOSINOFILOS"] = line.split()[1]
        elif "BAS#-BASOFILOS" in line:
            data["BAS#-BASOFILOS"] = line.split()[1]
        elif "RDW-CV" in line:
            data["RDW-CV"] = line.split()[1]
        elif "RDW-SD" in line:
            data["RDW-SD"] = line.split()[1]
        elif "PLT-PLAQUETAS" in line:
            data["PLT-PLAQUETAS"] = line.split()[1]
        elif "MPV" in line:
            data["MPV"] = line.split()[1]
        elif "PDW" in line:
            data["PDW"] = line.split()[1]
    return data

# Iterar sobre cada página del PDF y extraer datos
row_num = 2
for page_num in range(len(pdf)):
    page = pdf.load_page(page_num)
    text = page.get_text("text")  # Extraer el texto
    data = extract_data(text)  # Extraer datos relevantes
    if data:
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num, value=data.get(header, ""))
        row_num += 1

# Guardar el archivo Excel
output_file = 'resultados_organizado.xlsx'
workbook.save(output_file)
print(f"Archivo Excel generado: {output_file}")
