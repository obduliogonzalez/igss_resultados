import fitz  # PyMuPDF
import openpyxl

# Cargar el archivo PDF
pdf_document = 'C:/Users/nestor.gonzalez/Documents/GitHub/igss_resultados/saas_archivo_filtrado'
pdf = fitz.open(pdf_document)

# Crear un nuevo archivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Encabezados para la tabla
headers = [
    "ID", "Sexo", "RBC-ERITROCITOS", "HGB-HEMOGLOBINA", "HCT-HEMATOCRITO", "MCV", "MCH", "MCHC", 
    "WBC-LEUCOCITOS", "LYM%-LINFOCITOS", "NEU%-NEUTROFILOS", "MON%-MONOCITOS", 
    "EOS%-EOSINOFILOS", "BAS%-BASOFILOS", "LYN#-LINFOCITOS", "NEU#-NEUTROFILOS", 
    "MON#-MONOCITOS", "EOS#-EOSINOFILOS", "BAS#-BASOFILOS", "RDW-CV", "RDW-SD", 
    "PLT-PLAQUETAS", "MPV", "PDW"
]

# Escribir encabezados en Excel
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Función para buscar y extraer datos relevantes usando palabras y posiciones
def extract_data_by_words(page):
    """
    Extrae datos relevantes basados en palabras y posiciones horizontales.
    """
    words = page.get_text("words")  # Extraer todas las palabras con sus posiciones
    data = {key: "" for key in headers}  # Inicializar el diccionario de datos

    for word in words:
        x0, y0, x1, y1, text, block_no, line_no, word_no = word
        # Filtrar los datos según las etiquetas clave
        if "ID:" in text:
            data["ID"] = text.split(":")[-1]
        elif "Sexo:" in text:
            data["Sexo"] = text.split(":")[-1]
        elif text in headers:
            # Buscar el valor en la misma línea pero más a la derecha
            next_word = next((w[4] for w in words if w[1] == y0 and w[0] > x1), None)
            if next_word:
                data[text] = next_word

    return data

# Iterar sobre cada página del PDF y extraer datos
row_num = 2
for page_num in range(len(pdf)):
    page = pdf.load_page(page_num)
    data = extract_data_by_words(page)  # Extraer datos relevantes
    if any(data.values()):  # Verificar que hay datos válidos
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num, value=data.get(header, ""))
        row_num += 1

# Guardar el archivo Excel
output_file = 'resultados_procesados.xlsx'
workbook.save(output_file)
print(f"Archivo Excel generado: {output_file}")